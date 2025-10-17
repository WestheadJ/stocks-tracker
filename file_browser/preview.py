import os
from openpyxl import load_workbook


class SelectionMeta:
    def __init__(self, path):
        self.path = path
        self.size = None
        self.rows = None
        self.cols = None

        if os.path.isfile(path) and path.lower().endswith(".xlsx"):
            try:
                self.size = os.path.getsize(path)
            except Exception:
                pass

            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                self.rows = ws.max_row
                self.cols = ws.max_column
                wb.close()  # ✅ always close
            except Exception:
                pass
